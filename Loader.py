'''
Loads post-OCR PDFs into an Excel sheet using OpenPyXL

James Sy
July 27, 2022
'''

try:
    from OCR import Reader
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    from os.path import exists
    from plistlib import InvalidFileException

except ImportError as i:
    print(i.msg)

class Loader:
    '''
    Loads information pulled from documents into a new or existing Excel spreadsheet
    while also formatting the information and sheets

    Class attrs:

    att COLUMN_WIDTH: a constant specifying the width of the title columns
    inv: COLUMN_WIDTH is an int

    att reader: this Loader's associated Reader object
    inv: reader is an object of class Reader

    att excelname: the prefix of the Excel filename to be saved, 'newsheet' by default
    inv: excelname is a string

    att wstitle: the name of the saved spreadsheet sheet, 'Sheet' by default
    inv: wstitle is a string

    att wb: the active Workbook
    inv: wb is a Workbook or None
    '''

    COLUMN_WIDTH = 77

    def setexcelname(self, input): # assert isalpha, prevent illegal characters
        '''
        Setter for excelname
        Param: input must be a string of Windows-legal characters
        '''
        self.excelname = input

    def setwstitle(self, input):
        '''
        Setter for wstitle
        Param: input must be a string
        '''
        self.wstitle = input

    def getreader(self):
        '''
        Returns associated Reader object
        '''
        return self.reader

    def __init__(self):
        '''
        Initializer
        '''
        self.reader = Reader()
        self.excelname = 'newsheet' # default
        self.wstitle = 'Sheet' # default
        self.wb = None

    def excelinit(self, header1, header2, header3):
        '''
        For new workbook, initializes a new spreadsheet and populates it with column titles, including the user-inputted search queries
        '''

        headers = [
            'FILE PATH',
            'FIRST PARAGRAPH',
            header1.upper(),
            header2.upper(),
            header3.upper()
        ]

        self.inp1_title = header1.upper()
        self.inp2_title = header2.upper()
        self.inp3_title = header3.upper()

        try:
            self.wb = Workbook()
        except OSError as e:
            print(e.errno)
        self.sheet = self.wb.active
        self.sheet.title = self.wstitle

        for header_number in range(1, len(headers)+1):
            self.sheet.cell(column=header_number, row=1, value=headers[header_number-1])
            letter = get_column_letter(header_number)
            self.sheet.column_dimensions[letter].width = self.COLUMN_WIDTH

    def excelload(self):
        '''
        Calls reader's convert() method and loads this data into sheet within wb
        '''

        self.reader.convert()

        if self.reader.ret != []:
            for docdata in self.reader.ret:
                row_to_use = self.sheet.max_row + 1

                self.sheet.cell(column=1, row=row_to_use, value=docdata.getfile()) # A: filename
                self.sheet['A{}'.format(str(row_to_use))].alignment = Alignment(wrapText=True)

                self.sheet.cell(column=2, row=row_to_use, value=docdata.para()) # B: first paragraph
                self.sheet['B{}'.format(str(row_to_use))].alignment = Alignment(wrapText=True)
                self.sheet.column_dimensions['B'].width = self.COLUMN_WIDTH

                self.sheet.cell(column=3, row=row_to_use, value=docdata.input_1()) # C: first input
                self.sheet['C{}'.format(str(row_to_use))].alignment = Alignment(wrapText=True)
                self.sheet.column_dimensions['C'].width = self.COLUMN_WIDTH

                self.sheet.cell(column=4, row=row_to_use, value=docdata.input_2()) # D: second input
                self.sheet['D{}'.format(str(row_to_use))].alignment = Alignment(wrapText=True)
                self.sheet.column_dimensions['D'].width = self.COLUMN_WIDTH

                self.sheet.cell(column=5, row=row_to_use, value=docdata.input_3()) # E: third input
                self.sheet['E{}'.format(str(row_to_use))].alignment = Alignment(wrapText=True)
                self.sheet.column_dimensions['E'].width = self.COLUMN_WIDTH

    def savewb(self):
        '''
        If new == None, saves a new file, otherwise saves the existing Excel spreadsheet
        '''

        dest = self.excelname + '.xlsx'
        if not exists(dest):
            self.wb.save(self.reader.getdir() + '/' + dest)
        else:
            print('This file name already exists!')
        self.clear()

    def clear(self):
        '''
        Clears Loader fields and calls associated Reader object's clear() method
        '''
        self.getreader().clear()
        self.wb = None